#!/usr/bin/python3
import openpyxl
from openpyxl.styles import colors, PatternFill
from pathlib import Path

import re

greenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

import shutil
source = r'' # path to source xl-file
target = r'' # path to report xl-file
shutil.copyfile(source, target)

report = Path('') # path to report xl-file

wb = openpyxl.load_workbook(report, data_only=True)
summary = wb.create_sheet('Summary')
sheets = wb.sheetnames
sheets.sort()

wb['Summary']['A1'].value = 'Service Name'
wb['Summary']['B1'].value = 'Incident summary duration'

# service counter
services = []
for i in range(len(sheets)):
    ws = wb[sheets[i]]
    services.append(ws.max_row)


# collect values
for s in range(1, max(services)+1):
    service = []
    svcget = []
    for i in range(len(sheets)):
        ws = wb[sheets[i]]
        row = ws[s]
        for r in row:
            if type(r.value) is int:
                rc = r.coordinate
                rcnext = rc.replace('B', 'C')
                service.append(r.value)
            elif type(r.value) is str:
                svcname = re.findall(r'service ', r.value, re.I)
                if not svcname:
                    pass
                else:
                    rcget = r.coordinate
                    rcgetnext = rcget.replace('A', 'B')
                    svcget.append(r.value)
            else:
                pass

# summary content
    svcget = list(filter(None, list(set(svcget))))
    if not svcget:
        pass
    else:
        if not service:
            pass
        else:
            wb['Summary'][rcget] = str(svcget)
            wb['Summary'][rcgetnext] = sum(service)

# changes content
    if not service:
        pass
    else:
        for i in range(len(service)):
            sname = wb[sheets[0]]
            sname[rcnext] = '00'
            try:
                check = (service[i+1] - service[i])
                if check < 0:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
                    sname[rcnext].fill = greenFill
                elif check > 0:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
                    sname[rcnext].fill = redFill
                else:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
            except Exception as e:
                pass

#
wb.save(report)
