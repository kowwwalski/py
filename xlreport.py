#!/usr/bin/python3
import openpyxl
from openpyxl.styles import colors, PatternFill
from pathlib import Path

greenFill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
redFill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

import shutil
source = r'' # path to source xl-file
target = r'' # path to report xl-file
shutil.copyfile(source, target)

report = Path('') # path to report xl-file

wb = openpyxl.load_workbook(report, data_only=True)
sheets = wb.sheetnames
sheets.sort()

# service counter
services = []
for i in range(len(sheets)):
    ws = wb[sheets[i]]
    services.append(ws.max_row)

# collect values
for s in range(1, max(services)+1):
    service = []
    for i in range(len(sheets)):
        ws = wb[sheets[i]]
        row = ws[s]
        for r in row:
            if type(r.value) is not int:
                pass
            else:
                rc = r.coordinate
                rcnext = rc.replace('B', 'C') # column letter replace
                service.append(r.value)
    if not service:
        pass
    else:
        for i in range(len(service)):
            sname = wb[sheets[0]]
            sname[rcnext] = '00' # null for the 1st sheet
            try:
                check = (service[i+1] - service[i])
                if check < 0:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
                    sname[rcnext].fill = greenFill
                elif check > 0:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
                    sname[rcnext].fill = redFill
                else:
                    sname = wb[sheets[i+1]]
                    sname[rcnext] = check
            except Exception as e:
                pass
#
wb.save(report)
